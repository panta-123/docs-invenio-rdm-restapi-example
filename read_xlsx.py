import pandas as pd
from openpyxl import load_workbook
import json
from urllib.parse import unquote

file_path = "/Users/panta/Downloads/12GeV-proposals.xlsx"




df = pd.read_excel(file_path, skiprows=12)
#print(df.head(3))
# Initialize variables to store extracted data
proposal_spokesperson = {}


wb = load_workbook(file_path)
sheets = wb.sheetnames
ws = wb[sheets[0]]
"""
previous_proposal = None
for index, row in df.iterrows():
    proposal = row['Proposal']
    hall = row['Hall']
    title = row['Title']
    beam_days = row['Beam days']
    rating = row['Rating']
    pac_number = row['PAC']
if pd.notna(proposal):
    spokesperson = row['Spokesperson']
    institution = row['Institutions']
    doc = row['DOC']
    if pd.notna(previous_proposal):
    try:
        cell = ws.cell(row=index + 14, column=10)
        if cell.hyperlink:
            doc_hyperlink = cell.hyperlink.target
    except Exception as e:
        print(f"Error accessing cell: {e}")
"""


# Initialize variables to track previous filled values
previous_proposal = None
previous_spokesperson = None
previous_institution = None
previous_doc = None
previous_hall = None
previous_title = None
previous_beam_days = None
previous_rating = None
previous_pac_number = None

extracted_data_list = []

# Initialize variables to store data for the current proposal
current_proposal_data = {}

# Iterate through rows and extract data
for index, row in df.iterrows():
    proposal = row['Proposal']
    spokesperson = row['Spokesperson']
    institution = row['Institutions']
    doc = row['DOC']
    hall = row['Hall']
    title = row['Title']
    beam_days = row['Beam days']
    rating = row['Rating']
    pac_number = row['PAC']
    status = row['Status']
    # Check if the proposal value is not empty
    if pd.notna(proposal):
        try:
            cell = ws.cell(row=index+14, column=1)
            if cell.hyperlink:
                pac_hyperlink = cell.hyperlink.target
        except Exception as e:
            print(f"Error accessing cell: {e}")
        if current_proposal_data:
            extracted_data_list.append(current_proposal_data)
        
        # Update previous filled values with current row's values
        current_proposal_data = {
            'Proposal': proposal,
            #'Spokesperson': str(spokesperson),
            #'Institution': str(institution),
            #'people': [{'Name': spokesperson, 'Institution': institution}],
            'people': [{'spokes_person': str(spokesperson).rstrip("*") if str(spokesperson).endswith('*') else None,
                        'contact_person': str(spokesperson) if not str(spokesperson).endswith('*') else None,
                        'Institution': institution}],
            'Doc': str(doc),
            'Hall': hall,
            'Title': title,
            'Beam days': beam_days,
            'Rating': rating,
            'PAC': pac_number,
            'status': status
        }
        if pac_hyperlink:
            lss = pac_hyperlink.split('/')
            actual_hyperlink ="https://misportal.jlab.org/mis/physics/experiments/%s" %lss[-1]
            unquoted_hyperlink = unquote(actual_hyperlink)
            current_proposal_data['pac_hyperlink'] = unquoted_hyperlink
        
    else:
        #current_proposal_data['Spokesperson'] += f"\n{spokesperson}" if pd.notna(spokesperson) else ""
        #current_proposal_data['Institution'] += f"\n{institution}" if pd.notna(institution) else ""
        x = current_proposal_data['people']
        #x.append({'Name': f"{spokesperson}" if pd.notna(spokesperson) else "", 'Institution': f"{institution}" if pd.notna(institution) else ""})
        ##current_proposal_data['Doc'] += f"\n{doc}" if pd.notna(doc) else ""
        x.append({'spokes_person': str(spokesperson).rstrip("*") if str(spokesperson).endswith('*') else None,
                                                          'contact_person': str(spokesperson) if not str(spokesperson).endswith('*') else None,
                                                          'Institution': institution})

if current_proposal_data:
    extracted_data_list.append(current_proposal_data)
#print("Extracted data:")
print(extracted_data_list[-1])

meta = [] #{ "custom_fields" : [] }
for prop in extracted_data_list:
    print(prop['Rating'])
    #if prop['Beam days']:
    mydict = {"metadata": {} }
    try:
        bd = int(prop['Beam days'])
    except Exception as e:
        bd = 0
    x = {
    "pac:pac_number": int(prop['PAC']),
    "pac:beam_days": bd,
    "pac:proposal_number": prop["Proposal"],
    "pac:pac_status": {
        "id": prop['status'],
        "title": {
        "en":  prop['status'],
        }
    },
    "pac:pacdb_url":prop['pac_hyperlink'],
    "pac:pac_rating": {
        "id": prop['Rating'],
        "title": {
        "en":  prop['Rating'],
        }
    },
    "rdm:division": [
        {
        "id": "EPH" + str(prop['Hall']),
        "title": {
            "en": "Experimental Physics Hall " +str(prop['Hall']) 
        }
        }
    ]
    }
    mydict["custom_fields"] = x
    mydict["metadata"]["title"] = prop['Title']
    mydict["metadata"]["creators"] = []
    for role in prop['people']:
        creatorDict = {}
        if role['spokes_person']:
            firstname = role['spokes_person'].split(" ", 1)[0]
            lastname =  role['spokes_person'].split(" ", 1)[1]
            creatorDict["person_or_org"] = {"type": "personal", "given_name": firstname,\
                                       "family_name": lastname}
            creatorDict["role"] = {"id": "projectleader", "title": {"en": "Project leader"}}
            if not (role["Institution"] == "nan" or pd.isna(role["Institution"])):
              creatorDict["affiliations"] = []
              creatorDict["affiliations"].append({"name": role["Institution"]})
        if role['contact_person']:
            if not (role['contact_person'] == 'nan' or role['contact_person'] == '.'):
              firstname = role['contact_person'].split(" ", 1)[0]
              lastname =  role['contact_person'].split(" ", 1)[1]
              creatorDict["person_or_org"] = {"type": "personal", "given_name": firstname,\
                                        "family_name": lastname}
              creatorDict["role"] = {"id": "contactperson", "title": {"en": "Contact person"}}
              if not (role["Institution"] == "nan" or pd.isna(role["Institution"])):
                creatorDict["affiliations"] = []
                creatorDict["affiliations"].append({"name": role["Institution"]})
        mydict["metadata"]["creators"].append(creatorDict)
    mydict["metadata"]["resource_type"]= {"id": "publication-proposal"}
    mydict["metadata"]["publication_date"]= "2024-03-28"
    mydict["metadata"]["rights"]= [{"icon": "cc-by-icon","id": "cc-by-4.0",
                  "props": {
                    "url": "https://creativecommons.org/licenses/by/4.0/legalcode",
                    "scheme": "spdx"
                  },
                  "title": {
                    "en": "Creative Commons Attribution 4.0 International"
                  },
                  "description": {
                    "en": "The Creative Commons Attribution license allows re-distribution and re-use of a licensed work on the condition that the creator is appropriately credited."
                  }
                }
              ]
    mydict["access"] =  {"files": "public", "record": "public", "embargo": {"active": False}}
    mydict["files"] = {"enabled": False}
    meta.append(mydict)
jsss = json.dumps(meta, ensure_ascii=False, indent=4)
print(meta[9])
with open('1.json', 'w', encoding='utf-8') as f:
    json.dump(meta, f, ensure_ascii=False, indent=4)
